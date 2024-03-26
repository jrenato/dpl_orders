'''
Export groups to excel files
'''
import os

from django.core.management.base import BaseCommand, CommandError
from django.conf import settings
from django.utils.translation import gettext as _
from django.utils import dates

from openpyxl import load_workbook
from openpyxl.styles import Border, Alignment, Side, PatternFill
from tqdm import tqdm

from product_groups.models import ProductGroup


class Command(BaseCommand):
    '''
    Export groups to excel files
    '''
    help = 'Export groups to excel files'

    debug = False
    template_file = None
    align_center = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(border_style='thin'), right=Side(border_style='thin'),
        top=Side(border_style='thin'), bottom=Side(border_style='thin')
    )
    fill = PatternFill(fill_type='solid', start_color='00DEE6EF')

    def add_arguments(self, parser):
        parser.add_argument('--debug', action='store_true', help='Enable debug mode')

    def handle(self, *args, **options):
        # Load EXPORT_PATH from settings
        export_path = settings.EXPORT_PATH
        export_path = os.path.join(export_path, _('Newsletters Sheets'))

        self.template_file = os.path.join(
            settings.DOCUMENT_TEMPLATES_PATH, 'product_group_template.xltx'
        )

        if options['debug']:
            self.debug = True

        # Check if the export path exists
        if not os.path.isdir(export_path):
            raise CommandError(f'The export path "{export_path}" does not exist')

        produt_groups = self.get_pending_product_groups()

        for product_group in tqdm(produt_groups, desc='Exporting groups'):
            self.export_product_group(product_group, export_path)


    def get_pending_product_groups(self):
        '''
        Get the product groups with status 'PE'
        '''
        return ProductGroup.objects.filter(status='PE')


    def export_product_group(self, product_group, export_path):
        '''
        Export the product group to excel
        '''

        # Check for subdir named year/month_name in export path
        month_name = dates.MONTHS[product_group.created.month]
        subdir = os.path.join(export_path, product_group.created.strftime(f'%Y/{month_name}'))
        if not os.path.isdir(subdir):
            os.makedirs(subdir)

        # Create the excel file
        wb = load_workbook(self.template_file)
        wb.template = False
        ws = wb['Template']

        ws['A1'].value = product_group.name
        if product_group.customer_limit_date:
            ws['A9'].value = f'{_("Limit Date for the Customer")}:' \
                f'{product_group.customer_limit_date}'

        line = 11

        for item in tqdm(product_group.group_items.all(), desc='Exporting items', leave=False):
            ws[f'A{line}'].value = item.product.supplier.short_name
            ws[f'A{line}'].alignment = self.align_center
            ws[f'A{line}'].border = self.border
            
            ws[f'B{line}'].value = item.product.release_date
            ws[f'B{line}'].number_format = 'dd/mm/yyyy'
            ws[f'B{line}'].alignment = self.align_center
            ws[f'B{line}'].border = self.border

            ws[f'C{line}'].value = item.product.sku
            ws[f'C{line}'].alignment = self.align_center
            ws[f'C{line}'].border = self.border

            ws[f'D{line}'].value = item.product.name
            #ws[f'D{line}'].alignment = self.align_center
            ws[f'D{line}'].border = self.border

            ws[f'E{line}'].value = item.product.price
            ws[f'E{line}'].number_format = '[$R$-416] #,##0.00;[$R$-416] #,##0.00-'
            ws[f'E{line}'].alignment = self.align_center
            ws[f'E{line}'].border = self.border

            ws[f'F{line}'].value = None
            ws[f'F{line}'].number_format = '0'
            ws[f'F{line}'].alignment = self.align_center
            ws[f'F{line}'].border = self.border

            # Set fill if line is an even number
            if line % 2 == 0:
                ws[f'A{line}'].fill = self.fill
                ws[f'B{line}'].fill = self.fill
                ws[f'C{line}'].fill = self.fill
                ws[f'D{line}'].fill = self.fill
                ws[f'E{line}'].fill = self.fill
                ws[f'F{line}'].fill = self.fill

            line += 1

        export_filename = f'{subdir}/{product_group.name}.xlsx'
        wb.save(export_filename)
